import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

project = Path('/home/ubuntu/jacque-pegue-monte-layout')
source = Path('/home/ubuntu/upload/Jacque_Pegue_e_Monte_Catalogo_Completo(2).xlsx')
previous_path = project / 'data/catalogo-importado.json'
updated_path = project / 'data/catalogo-xlsx-atualizado.json'
diff_path = project / 'data/catalogo-xlsx-diff.json'

def slugify(value):
    normalized = unicodedata.normalize('NFD', value)
    normalized = ''.join(char for char in normalized if unicodedata.category(char) != 'Mn')
    return re.sub(r'(^-+|-+$)', '', re.sub(r'[^a-z0-9]+', '-', normalized.lower()))

def categorize(name):
    normalized = slugify(name)
    if re.search(r'cha|panela|revelacao|baby|casamento|boho|borboletas|personalizado|dourado|retro|romantico|batizado|debutante|ano-novo|natal|formatura|guns-n-roses|aniversario', normalized):
        return 'Celebrações'
    if re.search(r'vasco|sao-paulo|flamengo|futebol', normalized):
        return 'Esportes'
    if re.search(r'boiadeira|safari|fazendinha|jardim', normalized):
        return 'Temáticos'
    return 'Infantil'

workbook = load_workbook(source, read_only=True, data_only=True)
sheet = workbook.active
records = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    name, _, price, image, image_title = row
    if not isinstance(name, str) or not isinstance(price, str) or not isinstance(image, str):
        continue
    if price.strip() != 'R$ 170' or not image.startswith(('http://', 'https://')):
        continue
    records.append({
        'name': name.strip(),
        'price': price.strip(),
        'imageSource': image.strip(),
        'imageTitle': image_title.strip() if isinstance(image_title, str) else '',
        'category': categorize(name.strip()),
        'slug': slugify(name.strip()),
    })

previous = json.loads(previous_path.read_text())
previous_by_slug = {item['slug']: item for item in previous}
updated_by_slug = {item['slug']: item for item in records}

added = [item for slug, item in updated_by_slug.items() if slug not in previous_by_slug]
removed = [item for slug, item in previous_by_slug.items() if slug not in updated_by_slug]
changed = [
    {
        'slug': slug,
        'name': item['name'],
        'changes': {
            key: {'before': previous_by_slug[slug].get(key), 'after': item.get(key)}
            for key in ('name', 'price', 'imageSource', 'category')
            if previous_by_slug[slug].get(key) != item.get(key)
        },
    }
    for slug, item in updated_by_slug.items()
    if slug in previous_by_slug and any(previous_by_slug[slug].get(key) != item.get(key) for key in ('name', 'price', 'imageSource', 'category'))
]

updated_path.write_text(json.dumps(records, ensure_ascii=False, indent=2) + '\n')
diff_path.write_text(json.dumps({
    'previous_count': len(previous),
    'updated_count': len(records),
    'added': added,
    'removed': removed,
    'changed': changed,
}, ensure_ascii=False, indent=2) + '\n')

print(json.dumps({
    'previous_count': len(previous),
    'updated_count': len(records),
    'added': [item['name'] for item in added],
    'removed': [item['name'] for item in removed],
    'changed': [item['name'] for item in changed],
}, ensure_ascii=False, indent=2))
